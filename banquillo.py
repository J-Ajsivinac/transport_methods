import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Tuple

def find_shortest_cycle(matrix: List[List[int]], start: Tuple[int, int]) -> List[Tuple[int, int]]:
    rows, cols = len(matrix), len(matrix[0])
    shortest_cycle = None
    shortest_length = float('inf')
    
    def is_valid(x: int, y: int) -> bool:
        return 0 <= x < rows and 0 <= y < cols
    
    def has_vertical_horizontal_moves(path: List[Tuple[int, int]]) -> bool:
        vertical = any(path[i][0] != path[i + 1][0] for i in range(len(path) - 1))
        horizontal = any(path[i][1] != path[i + 1][1] for i in range(len(path) - 1))
        return vertical and horizontal
    
    def dfs(x: int, y: int, path: List[Tuple[int, int]], last_dir: str):
        nonlocal shortest_cycle, shortest_length
        
        if len(path) > 2 and (x, y) == start:
            if len(path) < shortest_length and has_vertical_horizontal_moves(path):
                shortest_cycle = path.copy()
                shortest_length = len(path)
            return
        
        if len(path) >= shortest_length:
            return
        
        directions = [('up', -1, 0), ('left', 0, -1), ('down', 1, 0), ('right', 0, 1)]
        for dir_name, dx, dy in directions:
            if dir_name == last_dir:
                continue
            
            new_x, new_y = x + dx, y + dy
            while is_valid(new_x, new_y):
                if matrix[new_x][new_y] != 0 or (new_x, new_y) == start:
                    if (new_x, new_y) not in path or ((new_x, new_y) == start and len(path) > 2):
                        if (new_x, new_y) == start:
                            dfs(new_x, new_y, path, dir_name)
                        else:
                            path.append((new_x, new_y))
                            dfs(new_x, new_y, path, dir_name)
                            path.pop()
                    if (new_x, new_y) == start:
                        return
                new_x, new_y = new_x + dx, new_y + dy
    
    if matrix[start[0]][start[1]] != 0:
        return []
    
    dfs(start[0], start[1], [start], "")
    
    return shortest_cycle if shortest_cycle else []

def stepping_stone_method(cost_matrix: np.ndarray, solution_matrix: np.ndarray, supplies: List[int], demands: List[int], wb: openpyxl.Workbook) -> Tuple[np.ndarray, int]:
    rows, cols = cost_matrix.shape
    total_cost = np.sum(solution_matrix * cost_matrix)
    step = 1
    
    def calculate_path_cost(path: List[Tuple[int, int]]) -> int:
        path_cost = 0
        for idx, (i, j) in enumerate(path):
            if idx % 2 == 0:
                path_cost += cost_matrix[i, j]
            else:
                path_cost -= cost_matrix[i, j]
        return path_cost

    def write_step_to_excel(step: int, solution: np.ndarray, all_paths: List[List[Tuple[int, int]]], all_costs: List[int], best_path: List[Tuple[int, int]]):
        ws = wb.create_sheet(title=f"Paso {step}")
        
        # Escribir la solución actual
        ws.cell(row=1, column=1, value="Solución actual:")
        for i in range(rows):
            for j in range(cols):
                ws.cell(row=i+2, column=j+1, value=solution[i, j])

        # Escribir todos los circuitos y sus costos
        start_row = rows + 4
        for idx, (path, cost) in enumerate(zip(all_paths, all_costs)):
            ws.cell(row=start_row + idx*3, column=1, value=f"Circuito {idx+1}:")
            ws.cell(row=start_row + idx*3 + 1, column=1, value=f"Camino: {path}")
            ws.cell(row=start_row + idx*3 + 2, column=1, value=f"Costo: {cost}")

            # Colorear las celdas del circuito en la solución, solo para el mejor ciclo
            if path == best_path:
                for k, (i, j) in enumerate(path):
                    cell = ws.cell(row=i+2, column=j+1)
                    if k % 2 == 0:
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # Calcular y escribir el costo total
        total_cost = np.sum(solution * cost_matrix)
        ws.cell(row=start_row + len(all_paths)*3 + 2, column=1, value=f"Costo total: {total_cost}")

    while True:
        all_paths = []
        all_costs = []
        best_path = None
        best_cost = float('inf')
        improvement_found = False
        
        for i in range(rows):
            for j in range(cols):
                if solution_matrix[i, j] == 0:
                    path = find_shortest_cycle(solution_matrix.tolist(), (i, j))
                    if path:
                        path_cost = calculate_path_cost(path)
                        all_paths.append(path)
                        all_costs.append(path_cost)
                        if path_cost < 0 and path_cost < best_cost:
                            best_cost = path_cost
                            best_path = path
                            improvement_found = True  # Marcar que se encontró una mejora

        if best_path:
            min_quantity = min(solution_matrix[best_path[k][0], best_path[k][1]] for k in range(1, len(best_path), 2))

            # Escribir el paso actual en Excel antes de la actualización
            write_step_to_excel(step, solution_matrix, all_paths, all_costs, best_path)

            for k, (x, y) in enumerate(best_path):
                if k % 2 == 0:
                    solution_matrix[x, y] += min_quantity
                else:
                    solution_matrix[x, y] -= min_quantity
            
            step += 1
        else:
            # Si no se encontró ninguna mejora, salir del bucle
            write_step_to_excel(step, solution_matrix, all_paths, all_costs, best_path)
            break

    return solution_matrix, np.sum(solution_matrix * cost_matrix)

def read_excel_data(filename: str) -> Tuple[np.ndarray, np.ndarray, List[int], List[int]]:
    wb = openpyxl.load_workbook(filename)
    
    # Leer matriz de costos
    ws_costs = wb['Costos']
    cost_matrix = np.array([[cell for cell in row[1:]] for row in ws_costs.iter_rows(min_row=2, values_only=True)])
    print("Dimensiones de la matriz de costos:", cost_matrix.shape)
    
    # Leer ofertas y demandas
    ws_supply_demand = wb['OfertaDemanda']
    
    # Imprimir información de depuración
    print("Contenido de la hoja OfertaDemanda:")
    for row in ws_supply_demand.iter_rows(values_only=True):
        print(row)
    
    supplies = []
    demands = []
    for row in ws_supply_demand.iter_rows(min_row=2, values_only=True):
        if len(row) >= 2:
            supplies.append(row[1])
        if len(row) >= 4:
            demands.append(row[3])
    
    print("Ofertas:", supplies)
    print("Demandas:", demands)
    
    # Leer solución inicial
    ws_solution = wb['SolucionInicial']
    solution_matrix = np.array([[cell for cell in row[1:]] for row in ws_solution.iter_rows(min_row=2, values_only=True)])
    print("Dimensiones de la matriz de solución inicial:", solution_matrix.shape)
    
    return cost_matrix, solution_matrix, supplies, demands
# ... [El resto del código permanece igual]

def main():
    filename = 'datos_banquillo.xlsx'
    cost_matrix, solution_matrix, supplies, demands = read_excel_data(filename)
    
    print("Matriz de costos:")
    print(cost_matrix)
    print("\nMatriz de solución inicial:")
    print(solution_matrix)
    print("\nOfertas:", supplies)
    print("Demandas:", demands)
    
    wb = openpyxl.Workbook()
    optimized_solution, optimized_cost = stepping_stone_method(cost_matrix, solution_matrix, supplies, demands, wb)
    
    # Escribir la solución final
    ws_final = wb.create_sheet(title="Solución Final")
    ws_final.cell(row=1, column=1, value="Solución optimizada:")
    for i in range(len(optimized_solution)):
        for j in range(len(optimized_solution[0])):
            ws_final.cell(row=i+2, column=j+1, value=optimized_solution[i, j])
    ws_final.cell(row=len(optimized_solution)+3, column=1, value=f"Costo optimizado: {optimized_cost}")
    
    # Guardar el archivo Excel con los resultados
    wb.save('resultados_stepping_stone.xlsx')
    print("Resultados guardados en 'resultados_stepping_stone.xlsx'")

if __name__ == "__main__":
    main()