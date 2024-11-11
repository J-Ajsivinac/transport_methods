import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def leer_datos_excel(archivo):
    wb = openpyxl.load_workbook(archivo)
    ws = wb.active
    
    datos = [[cell.value for cell in row] for row in ws.iter_rows()]
    
    ultima_fila = max(i for i, row in enumerate(datos) if any(cell is not None for cell in row))
    ultima_columna = max(j for row in datos for j, cell in enumerate(row) if cell is not None)
    
    oferta = [row[ultima_columna] for row in datos[1:ultima_fila] if row[ultima_columna] is not None]
    demanda = [datos[ultima_fila][j] for j in range(1, ultima_columna) if datos[ultima_fila][j] is not None]
    costos = [[row[j] for j in range(1, ultima_columna) if row[j] is not None] for row in datos[1:ultima_fila]]
    
    print("Oferta:", oferta)
    print("Demanda:", demanda)
    print("Costos:", costos)
    
    oferta = [float(val) if val is not None else 0 for val in oferta]
    demanda = [float(val) if val is not None else 0 for val in demanda]
    costos = [[float(val) if val is not None else 0 for val in row] for row in costos]
    
    max_len = max(len(row) for row in costos)
    costos = [row + [0] * (max_len - len(row)) for row in costos]
    
    return oferta, demanda, costos, wb

def esquina_noroeste(oferta, demanda, wb):
    if "solucion_noroeste" not in wb.sheetnames:
        ws = wb.create_sheet("solucion_noroeste")
    else:
        ws = wb["solucion_noroeste"]
        ws.delete_rows(1, ws.max_row)  # Limpiar la hoja si ya existe
    m, n = len(oferta), len(demanda)
    x = np.zeros((m, n))
    i, j = 0, 0
    paso = 1
    fila_actual = 1
    
    while i < m and j < n:
        cantidad = min(oferta[i], demanda[j])
        x[i, j] = cantidad
        oferta[i] -= cantidad
        demanda[j] -= cantidad
        
        fila_actual = escribir_paso_excel(ws, x, oferta, demanda, i, j, cantidad, paso, fila_actual)
        paso += 1
        
        if oferta[i] == 0:
            i += 1
        if demanda[j] == 0:
            j += 1
    
    return x

def escribir_paso_excel(ws, x, oferta, demanda, i, j, cantidad, paso, fila_actual):
    ws.cell(row=fila_actual, column=1, value=f"Paso {paso}")
    fila_actual += 1
    
    for row in range(len(x)):
        for col in range(len(x[0])):
            cell = ws.cell(row=fila_actual+row, column=col+2, value=x[row][col])
            if row == i and col == j:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for row in range(len(oferta)):
        ws.cell(row=fila_actual+row, column=len(x[0])+2, value=oferta[row])
    
    for col in range(len(demanda)):
        ws.cell(row=fila_actual+len(x), column=col+2, value=demanda[col])
    
    ws.cell(row=fila_actual+len(x)+1, column=1, value=f"Asignación: ({i+1}, {j+1}) = {cantidad}")
    
    for row in ws[f"B{fila_actual}:{get_column_letter(len(x[0])+2)}{fila_actual+len(x)}"]:
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
    
    fila_actual += len(x) + 3  # Agregamos un espacio extra entre pasos
    return fila_actual

def agregar_variables_ficticias(oferta, demanda, costos):
    total_oferta = sum(oferta)
    total_demanda = sum(demanda)
    
    print(f"Total oferta: {total_oferta}")
    print(f"Total demanda: {total_demanda}")
    
    if total_oferta > total_demanda:
        demanda.append(total_oferta - total_demanda)
        for fila in costos:
            fila.append(0)
    elif total_demanda > total_oferta:
        oferta.append(total_demanda - total_oferta)
        costos.append([0] * len(demanda))
    
    return oferta, demanda, costos

def calcular_costo_total(solucion, costos):
    return np.sum(solucion * np.array(costos))

def resolver_problema_transporte(archivo_excel):
    oferta, demanda, costos, wb = leer_datos_excel(archivo_excel)
    
    if not oferta or not demanda or not costos:
        print("Error: Datos incompletos en el archivo Excel.")
        return None, None
    
    oferta, demanda, costos = agregar_variables_ficticias(oferta, demanda, costos)
    solucion = esquina_noroeste(oferta.copy(), demanda.copy(), wb)
    costo_total = calcular_costo_total(solucion, costos)
    
    ws = wb["solucion_noroeste"]
    ultima_fila = ws.max_row + 2
    ws.cell(row=ultima_fila, column=1, value="Solución Final:")
    for i in range(len(solucion)):
        for j in range(len(solucion[0])):
            ws.cell(row=ultima_fila+i+1, column=j+2, value=solucion[i][j])
    
    ws.cell(row=ultima_fila+len(solucion)+2, column=1, value=f"Costo total: {costo_total}")
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(archivo_excel)
    return solucion, costo_total

# Ejemplo de uso
archivo_excel = "tabla_transporte.xlsx"
solucion, costo_total = resolver_problema_transporte(archivo_excel)
if solucion is not None and costo_total is not None:
    print("Solución guardada en el archivo Excel.")
    print("Costo total:", costo_total)
else:
    print("No se pudo resolver el problema debido a datos incompletos o incorrectos.")