import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
from typing import List, Tuple
import os

def find_shortest_cycle(matrix: List[List[int]], start: Tuple[int, int]) -> List[Tuple[int, int]]:
    rows, cols = len(matrix), len(matrix[0])
    shortest_cycle = None
    shortest_length = float('inf')
    
    def is_valid(x: int, y: int) -> bool:
        return 0 <= x < rows and 0 <= y < cols
    
    def has_vertical_horizontal_moves(path: List[Tuple[int, int]]) -> bool:
        vertical = any(path[i][0] != path[i + 1][0] for i in range(len(path) - 1))
        horizontal = any(path[i][1] != path[i + 1][1] for i in range(len(path) - 1))
        return vertical and horizontal
    
    def dfs(x: int, y: int, path: List[Tuple[int, int]], last_dir: str):
        nonlocal shortest_cycle, shortest_length
        
        if len(path) > 2 and (x, y) == start:
            if len(path) < shortest_length and has_vertical_horizontal_moves(path):
                shortest_cycle = path.copy() 
                shortest_length = len(path)
            return
        
        if len(path) >= shortest_length:
            return
        
        directions = [('up', -1, 0), ('left', 0, -1), ('down', 1, 0), ('right', 0, 1)]
        for dir_name, dx, dy in directions:
            if dir_name == last_dir:
                continue
            
            new_x, new_y = x + dx, y + dy
            while is_valid(new_x, new_y):
                if matrix[new_x][new_y] != 0 or (new_x, new_y) == start:
                    if (new_x, new_y) not in path or ((new_x, new_y) == start and len(path) > 2):
                        if (new_x, new_y) == start:
                            dfs(new_x, new_y, path, dir_name)
                        else:
                            path.append((new_x, new_y))
                            dfs(new_x, new_y, path, dir_name)
                            path.pop()
                    if (new_x, new_y) == start:
                        return  
                new_x, new_y = new_x + dx, new_y + dy
    
    if matrix[start[0]][start[1]] != 0:
        return []  
    
    dfs(start[0], start[1], [start], "")
    
    return shortest_cycle if shortest_cycle else []


def modi_method(cost_matrix: np.ndarray, supply: List[int], demand: List[int], initial_solution: np.ndarray, wb: openpyxl.Workbook, sheet_name: str) -> np.ndarray:
    rows, cols = cost_matrix.shape
    u = np.full(rows, None)
    v = np.full(cols, None)
    solution = initial_solution.copy()
    u[0] = 0
    
    ws = wb.create_sheet(sheet_name)
    step = 1

    solution = solution.astype(float)

    while True:
        while None in u or None in v:
            for i in range(rows):
                for j in range(cols):
                    if solution[i, j] > 0:
                        if u[i] is not None and v[j] is None:
                            v[j] = cost_matrix[i, j] - u[i]
                        elif u[i] is None and v[j] is not None:
                            u[i] = cost_matrix[i, j] - v[j]

        reduced_costs = np.zeros_like(cost_matrix, dtype=float)
        for i in range(rows):
            for j in range(cols):
                if solution[i, j] == 0:
                    reduced_costs[i, j] = cost_matrix[i, j] - (u[i] + v[j])

        write_step_to_excel(ws, step, solution, u, v, reduced_costs)
        step += 1

        if np.all(reduced_costs >= 0):
            ws.cell(row=step*15, column=1, value="La solución es óptima.")
            step += 1
            total_cost = np.sum(solution * cost_matrix)
            ws.cell(row=step*15 + 1, column=1, value="Costo total:")
            ws.cell(row=step*15 + 2, column=1, value=total_cost)
            break

        entering_variable = np.unravel_index(np.argmin(reduced_costs), reduced_costs.shape)
        cycle = find_shortest_cycle(solution.tolist(), entering_variable)
        
        if cycle:
            min_val = min(solution[cycle[i]] for i in range(1, len(cycle), 2) if solution[cycle[i]] > 0)

            for i, (x, y) in enumerate(cycle):
                if i % 2 == 0:
                    solution[x, y] += min_val
                else:
                    solution[x, y] -= min_val

            u.fill(None)
            v.fill(None)
            u[0] = 0

        write_step_to_excel(ws, step, solution, u, v, reduced_costs, cycle=cycle)
    
    return solution

def write_step_to_excel(ws, step, solution, u, v, reduced_costs, cycle=None):
    start_row = (step - 1) * 15 + 1  # Espaciado entre pasos
    ws.cell(row=start_row, column=1, value=f"Paso {step}")
    
    ws.cell(row=start_row + 1, column=1, value="Solución actual:")
    for i in range(solution.shape[0]):
        for j in range(solution.shape[1]):
            cell = ws.cell(row=start_row + 2 + i, column=j + 1, value=solution[i, j])
            

    if cycle:
        cycle_vertices = " -> ".join(f"({x+1}, {y+1})" for (x, y) in cycle)
        ws.cell(row=start_row + 1, column=solution.shape[1] + 4, value=f"Circuito usado: {cycle_vertices}")
        
        for i in range(len(cycle)):
            x, y = cycle[i]
            ws.cell(row=start_row + 2 + x, column=y + 1).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    ws.cell(row=start_row + 1, column=solution.shape[1] + 2, value="Multiplicadores u:")
    for i, val in enumerate(u):
        ws.cell(row=start_row + 2 + i, column=solution.shape[1] + 2, value=f"u_{i+1}: {val}")

    ws.cell(row=start_row + 1, column=solution.shape[1] + 3, value="Multiplicadores v:")
    for j, val in enumerate(v):
        ws.cell(row=start_row + 2, column=solution.shape[1] + 3 + j, value=f"v_{j+1}: {val}")

    formula_start_row = start_row + solution.shape[0] + 4

    ws.cell(row=formula_start_row, column=1, value="Fórmulas X_ij (solo para 0):")
    for i in range(solution.shape[0]):
        for j in range(solution.shape[1]):
            if solution[i, j] == 0:
                if(reduced_costs is not None):
                    formula = f"X_{i+1}{j+1} = C_{i+1}{j+1} - (u_{i+1} + v_{j+1}) = {reduced_costs[i, j]}"
                else:
                    formula = f"X_{i+1}{j+1} = C_{i+1}{j+1} - (u_{i+1} + v_{j+1})"
                ws.cell(row=formula_start_row + 1 + i, column=j + 1, value=formula)


def read_excel_data(file_path: str):
    wb = openpyxl.load_workbook(file_path)
    
    ws_costs = wb["Costos"]
    cost_matrix = [[float(cell.value) if cell.value is not None and cell.value != '' else 0 
                for cell in row[1:]]  # Comenzar desde la segunda columna
                for row in ws_costs.iter_rows(min_row=2, max_col=ws_costs.max_column, max_row=ws_costs.max_row)]  # Comenzar desde la segunda fila
    cost_matrix = np.array(cost_matrix)

    ws_supply_demand = wb["Oferta y Demanda"]
    supply = []
    demand = []
    for row in ws_supply_demand.iter_rows(min_row=2, values_only=True):
        if row[1] is not None and row[1] != '':
            supply.append(float(row[1]))
        if row[2] is not None and row[2] != '':
            demand.append(float(row[2]))

    ws_initial = wb["Solucion Inicial"]
    initial_solution = [[float(cell.value) if cell.value is not None and cell.value != '' else 0 
                    for cell in row[1:]]  # Comenzar desde la segunda columna
                    for row in ws_initial.iter_rows(min_row=2, max_col=ws_initial.max_column, max_row=ws_initial.max_row)]
    initial_solution = np.array(initial_solution)

    return wb, cost_matrix, supply, demand, initial_solution

path_v = os.path.abspath(__file__)
file_path = os.path.join(os.path.dirname(path_v), 'datos_transporte.xlsx')

wb, cost_matrix, supply, demand, initial_solution = read_excel_data(file_path)

print("Matriz de costos:")
print(cost_matrix)
print("\nOferta:")
print(supply)
print("\nDemanda:")
print(demand)
print("\nSolución inicial:")
print(initial_solution)

optimal_solution = modi_method(cost_matrix, supply, demand, initial_solution, wb, "MODI Steps")
print(optimal_solution)
archivo_salida = os.path.join(os.path.dirname(path_v), 'resultado_MODI.xlsx')
wb.save(archivo_salida)

print("Proceso completado. Revise el archivo 'resultados_modi.xlsx' para ver los pasos y resultados.")