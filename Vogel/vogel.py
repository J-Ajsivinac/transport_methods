import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill
import os
def vogel_approximation(supply, demand, costs, writer, sheet_name="Resultados"):
    original_supply = supply.copy()
    original_demand = demand.copy()
    original_costs = costs.copy()
    
    costs = costs.astype(float) 
    
    supply = supply.copy()
    demand = demand.copy()
    
    m, n = len(supply), len(demand)  # Dimensiones de la matriz
    
    df_original = pd.DataFrame(costs, columns=[f"Destino {i+1}" for i in range(n)], index=[f"Origen {i+1}" for i in range(m)])
    df_original["Oferta"] = supply
    demanda_df = pd.DataFrame([demand], columns=[f"Destino {i+1}" for i in range(n)])
    demanda_df.index = ["Demanda"]
    
    df_original.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=0)
    demanda_df.to_excel(writer, sheet_name=sheet_name, startrow=m+2, startcol=1)
    if sum(supply) > sum(demand):
        demand = np.append(demand, sum(supply) - sum(demand)) 
        costs = np.column_stack((costs, np.zeros(m)))  
    elif sum(supply) < sum(demand):
        supply = np.append(supply, sum(demand) - sum(supply))  
        costs = np.vstack((costs, np.zeros(n)))  
    
    m, n = len(supply), len(demand)
    x = np.zeros((m, n))  
    
    step = 1  
    
    while np.sum(supply) > 0 and np.sum(demand) > 0:
        row_penalties = np.zeros(m)
        col_penalties = np.zeros(n)
        
        for i in range(m):
            if supply[i] > 0:  
                sorted_costs = np.sort(costs[i, :])
                row_penalties[i] = sorted_costs[1] - sorted_costs[0] if len(sorted_costs) > 1 else 0
        
        for j in range(n):
            if demand[j] > 0:  
                sorted_costs = np.sort(costs[:, j])
                col_penalties[j] = sorted_costs[1] - sorted_costs[0] if len(sorted_costs) > 1 else 0
        
        penal_df_rows = pd.DataFrame({'Penalización Filas': row_penalties})  # Penalizaciones de filas
        penal_df_cols = pd.DataFrame({'Penalización Columnas': col_penalties})  # Penalizaciones de columnas
        
        penal_df_rows.to_excel(writer, sheet_name=sheet_name, startrow=(step*10), startcol=n+2)
        penal_df_cols.to_excel(writer, sheet_name=sheet_name, startrow=(step*10), startcol=n+4)

        if np.max(row_penalties) >= np.max(col_penalties):
            i = np.argmax(row_penalties)
            j = np.argmin(costs[i, :])
        else:
            j = np.argmax(col_penalties)
            i = np.argmin(costs[:, j])
        
        quantity = min(supply[i], demand[j])
        x[i, j] = quantity
        supply[i] -= quantity
        demand[j] -= quantity
        # Guardar el paso actual en Excel
        df_solution = pd.DataFrame(x, columns=[f"Destino {i+1}" for i in range(n)], index=[f"Origen {i+1}" for i in range(m)])
        df_solution["Oferta"] = supply
        demanda_df = pd.DataFrame([demand], columns=[f"Destino {i+1}" for i in range(n)])
        demanda_df.index = ["Demanda"]
        
        df_solution.to_excel(writer, sheet_name=sheet_name, startrow=(step*10), startcol=0)
        demanda_df.to_excel(writer, sheet_name=sheet_name, startrow=(step*10) + m + 2, startcol=1)
        
        if supply[i] == 0:
            costs[i, :] = np.inf  
        if demand[j] == 0:
            costs[:, j] = np.inf  

        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  
        if supply[i] == 0:
            for col in range(1, n+2):
                cell = worksheet.cell(row=(step*10) + i + 2, column=col)
                cell.fill = fill
        if demand[j] == 0:
            for row in range(2, m+2):
                cell = worksheet.cell(row=row + (step*10), column=j+2)
                cell.fill = fill
        
        step += 1
    return x, original_supply, original_demand, original_costs
# Leer los datos desde el archivo Excel
path_v = os.path.abspath(__file__)
file_path = os.path.join(os.path.dirname(path_v), 'tabla_transporte.xlsx')
# file_path = "tabla_transporte.xlsx"
df = pd.read_excel(file_path, sheet_name="Datos", index_col=0)
costs = df.iloc[:-1, :-1].values  # Matriz de costos
supply = df.iloc[:-1, -1].values  # Columna de suministro
demand = df.iloc[-1, :-1].values  # Fila de demanda


output_file = os.path.join(os.path.dirname(path_v), "resultados_vogel.xlsx")
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    solution, original_supply, original_demand, original_costs = vogel_approximation(supply, demand, costs, writer)

m, n = len(original_supply), len(original_demand)
total_cost = np.sum(solution[:m, :n] * original_costs)
print(f"\nCosto total: {total_cost}")